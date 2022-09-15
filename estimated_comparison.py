import openpyxl
import os
import datetime
import pyodbc
import exchangelib
import sys
import re

path = fr'D:\Daily Report'
# path = fr'D:\daily'
desks = {
    "BENI SUEF_DESK": {"producers": 0, "injectors": 0},
    "KARAMA_DESK": {"producers": 0, "injectors": 0},
    "QARUN_DESK": {"producers": 0, "injectors": 0},
    "WADY EL RAYAN_DESK": {"producers": 0, "injectors": 0},
    "EAST BAHARYA_DESK": {"producers": 0, "injectors": 0},
    "EAST BAHARYA EX03_DESK": {"producers": 0, "injectors": 0}
}


def main():

    startDateStr = '2022-08-01'
    endDateStr = '2022-09-01'
    print(startDateStr)
    print(endDateStr)

    startDate = datetime.datetime.strptime(startDateStr, '%Y-%m-%d')
    endDate = datetime.datetime.strptime(endDateStr, '%Y-%m-%d')

    excelData = dict()

    # while startDate != endDate:
    #     print(startDate)
    #     startDate += datetime.timedelta(days=1)

    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=QPCAVODB2K12;'
                          'Database=AVOCET_PRODUCTION;'
                          'Trusted_Connection=yes;'
                          'UID=sa;'
                          'PWD=A@qpc2017;')

    cursor = conn.cursor()
    cursor.execute(f'''SELECT D.[DATETIME], C.WELL, C.ITEM_NAME, C.START_DATETIME, C.END_DATETIME, C.[STATUS], C.[TYPE], ISNULL(T.PROD_OIL_VOL, 0) AS PROD_OIL_VOL, ISNULL(T.PROD_LIQ_VOL, 0) AS PROD_LIQ_VOL, ISNULL(DW.DURATION, 0) AS DURATION
    , dbo.INAME(DESK.FROM_ITEM_ID) AS DESK, ISNULL(R.WINJ_RATE, 0) AS WINJ_RATE, R.WAT_DISP
    ,TEST.START_DATETIME, TEST.ITEM_NAME, TEST.LIQ_VOL, TEST.BSW_VOL_FRAC * 100 AS BSW, TEST.OIL_VOL
    ,PUMP_INTK_PRESS
    FROM
    AVOCET_PRODUCTION.dbo.DATE_INFO AS D
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VI_COMPLETION_ALL_en_US AS C
    ON C.START_DATETIME <= D.[DATETIME]
    AND
    C.END_DATETIME > D.[DATETIME]
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VT_TOTALS_DAY_en_US AS T
    ON
    T.ITEM_ID = C.ITEM_ID
    AND
    T.START_DATETIME = D.[DATETIME]
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VT_WELL_READ_en_US AS R
    ON
    R.ITEM_ID = C.ITEM_ID
    AND
    R.START_DATETIME = D.[DATETIME]
    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VL_DESK_ITEM_en_US AS DESK
    ON
    DESK.TO_ITEM_ID = C.ITEM_ID

    LEFT JOIN
    AVOCET_PRODUCTION.dbo.VT_DOWNTIME_en_US AS DW
    ON
    DW.ITEM_ID = C.ITEM_ID
    AND
    CAST(DW.START_DATETIME AS DATE) = CAST(D.[DATETIME] AS DATE)
    LEFT JOIN
    VT_WELL_TEST_en_US AS TEST
    ON
    TEST.ITEM_ID = C.ITEM_ID
    AND
    TEST.START_DATETIME = (SELECT TOP 1 START_DATETIME FROM VT_WELL_TEST_en_US WHERE START_DATETIME <= D.[DATETIME] AND ITEM_ID = C.ITEM_ID AND VALID_TEST = 'True' ORDER BY START_DATETIME DESC)
    AND
    TEST.VALID_TEST = 'True'


    WHERE
    D.[DATETIME] >= '{startDateStr}'
    AND
    D.[DATETIME] < '{endDateStr}'
    AND
    D.DATE_TYPE = 'D'
    ORDER BY
    D.[DATETIME]''')
    avocetData = dict()
    for row in cursor:
        if row.DATETIME not in avocetData:
            avocetData[row.DATETIME] = dict()
            excelData[row.DATETIME] = getExcelForDay(
                row.DATETIME + datetime.timedelta(days=1))
        if not row.WELL:
            continue
        row.WELL = row.WELL.strip()
        if row.WELL in avocetData[row.DATETIME]:
            if avocetData[row.DATETIME][row.WELL]["rate"] > row.PROD_OIL_VOL or avocetData[row.DATETIME][row.WELL]["inj_rate"] > row.WINJ_RATE:
                continue
        avocetData[row.DATETIME][row.WELL] = {
            "status": row.STATUS,
            "hrs_online": 24 - (row.DURATION / 60 / 60),
            "rate": int(row.PROD_OIL_VOL),
            "bopd": row.OIL_VOL,
            "bsw": row.BSW,
            "bfpd": row.LIQ_VOL,
            "name": row.ITEM_NAME,
            "inj_rate": row.WINJ_RATE,
            "desk": row.DESK,
            "pIntake": row.PUMP_INTK_PRESS
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producers"
    ws.append(["Production Date",
               "Excel Name", "Avocet Name",
               "Excel Estimated Oil", "Avocet Estimated Oil",
               "Excel Hrs Online", "Avocet Hrs Online",
               "Excel Test BF", "Avocet Test BF",
               "Excel Test BSW", "Avocet Test BSW",
               "Excel Test BO", "Avocet Test BO",
               "Excel Status", "Avocet Status"])
    ws2 = wb.create_sheet("Injectors")
    ws2.append(["Production Date",
                "Excel Name", "Avocet Name",
                "Excel Inj Water", "Avocet Inj Water",
                "Excel Hrs Online", "Avocet Hrs Online",
                "Excel Status", "Avocet Status"])
    ws3 = wb.create_sheet("Sensors")
    ws3.append(["Production Date",
                "Excel Name", "Avocet Name",
                "Excel PI", "Avocet PI"
                ])

    for date in excelData:
        for well in excelData[date]:
            if "rate" in excelData[date][well]:
                if abs(int(avocetData[date][well]["rate"]) - int(excelData[date][well]["rate"])) > 2:
                    desks[avocetData[date][well]["desk"]]["producers"] += 1
                    ws.append([date.strftime('%Y-%m-%d'),
                               well, avocetData[date][well]["name"],
                               excelData[date][well]["rate"], avocetData[date][well]["rate"],
                               excelData[date][well]["hrs_online"], avocetData[date][well]["hrs_online"],
                               excelData[date][well]["bfpd"], avocetData[date][well]["bfpd"],
                               excelData[date][well]["bsw"], avocetData[date][well]["bsw"],
                               excelData[date][well]["bopd"], avocetData[date][well]["bopd"],
                               excelData[date][well]["status"], avocetData[date][well]["status"]
                               ])
                
                addToSheet = False
                if avocetData[date][well]["pIntake"] is not None:
                    try:
                        float(avocetData[date][well]["pIntake"])
                    except ValueError:
                        avocetData[date][well]["pIntake"] = None
                if excelData[date][well]["pIntake"] is not None:
                    try:
                        float(excelData[date][well]["pIntake"])
                    except ValueError:
                        excelData[date][well]["pIntake"] = None


                if avocetData[date][well]["pIntake"] is not None and excelData[date][well]["pIntake"] is None:
                    if int(round(avocetData[date][well]["pIntake"])) != 15:
                        print(avocetData[date][well]["name"])
                        print(type(avocetData[date][well]["pIntake"]))
                        print(round(avocetData[date][well]["pIntake"]) == 15)
                        print(avocetData[date][well]["pIntake"])
                        print(excelData[date][well]["pIntake"])

                        addToSheet = True
                elif avocetData[date][well]["pIntake"] is None and excelData[date][well]["pIntake"] is not None:
                    addToSheet = True


                if avocetData[date][well]["pIntake"] is not None and excelData[date][well]["pIntake"] is not None:
                    if abs(float(avocetData[date][well]["pIntake"]) - (float(excelData[date][well]["pIntake"]) + 14.70)) > 0:
                        addToSheet = True
                    
                if addToSheet == True:    
                    ws3.append([date.strftime('%Y-%m-%d'),
                               well, avocetData[date][well]["name"],
                               excelData[date][well]["pIntake"], avocetData[date][well]["pIntake"]
                               ])
                
            elif "inj_rate" in excelData[date][well]:

                if abs(int(avocetData[date][well]["inj_rate"]) - int(excelData[date][well]["inj_rate"])) > 1:
                    desks[avocetData[date][well]["desk"]]["injectors"] += 1
                    ws2.append([date.strftime('%Y-%m-%d'),
                                well, avocetData[date][well]["name"],
                                excelData[date][well]["inj_rate"], avocetData[date][well]["inj_rate"],
                                excelData[date][well]["hrs_online"], avocetData[date][well]["hrs_online"],
                                excelData[date][well]["status"], avocetData[date][well]["status"]
                                ])

    wb.save(f"D:\daily\Comparison from {startDateStr} to {endDateStr}.xlsx")

    table = "<table style='border: 1px solid black;'><tr><td style='border: 1px solid black;'>Desk</td><td style='border: 1px solid black;'>Producers</td><td style='border: 1px solid black;'>Injectors</td></tr>"
    for desk in desks:
        table += "<tr><td style='border: 1px solid black;'>" + desk + "</td><td style='border: 1px solid black;'>" + \
            str(desks[desk]["producers"]) + "</td><td style='border: 1px solid black;'>" + \
            str(desks[desk]["injectors"]) + "</td></tr>"

    table += "</table>"

    credentials = exchangelib.Credentials('mkarm', 'passwordHere')
    config = exchangelib.Configuration(
        server='webmail.qarun.net', credentials=credentials)

    account = exchangelib.Account(
        'mkarm@qarun.net', credentials=credentials, config=config)


    emailBody = f'''Gents,

                    Kindly find the attached file for the consistency check between Avocet and Excel daily reports for the period between {startDateStr} and {endDateStr}, production dates. Your assistance is highly appreciated to make the necessary corrections in Avocet to have consistent data. If any assistance is required from our side, please let us know.

                    Sincerely,
                    Mohamed
                '''


    message = exchangelib.Message(account=account,
                                  folder=account.sent,
                                  subject=f'Data Consistency check for {startDateStr} to {endDateStr}',
                                  body=exchangelib.HTMLBody(
                                      re.sub('[\n]', '<br>', emailBody) or f'Please find the attached file for the consistency check between Avocet and Excel for the period  {startDateStr} to {endDateStr}, production dates.<br><br>' + table),

                                  to_recipients=[exchangelib.Mailbox(
                                      email_address='mkarm@qarun.net'), exchangelib.Mailbox(
                                      email_address='mmabdelfattah@qarun.net')]
                                  )

    with open(fr"D:\daily\Comparison from {startDateStr} to {endDateStr}.xlsx", 'rb') as binary_file_content:
        my_other_file = exchangelib.FileAttachment(
            name=f"Comparison from {startDateStr} to {endDateStr}.xlsx", content=binary_file_content.read())

    message.attach(my_other_file)
    message.send_and_save()

    # for row in cursor:
    #     if row.DATETIME not in excelData:
    #         excelData[row.DATETIME] = getExcelForDay(row.DATETIME + datetime.timedelta(days=1))

    #     if row.WELL in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION' and row.STATUS == 'PRODUCING':
    #         if abs(int(row.PROD_OIL_VOL) - int(excelData[row.DATETIME][row.WELL]["bopd"] * excelData[row.DATETIME][row.WELL]["hrs_online"] / 24)) < 3:
    #             del excelData[row.DATETIME][row.WELL]
    #             continue
    #     elif row.WELL not in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION' and row.STATUS == 'PRODUCING':
    #         print(row.WELL + " " + row.ITEM_NAME + " not found in Excel")

    #     elif row.WELL in excelData[row.DATETIME] and row.TYPE == 'PRODUCTION':
    #         if excelData[row.DATETIME][row.WELL]["hrs_online"] > 0:
    #             if abs(int(row.PROD_OIL_VOL) - int(excelData[row.DATETIME][row.WELL]["bopd"] * excelData[row.DATETIME][row.WELL]["hrs_online"] / 24)) > 3:
    #                 print(row.PROD_OIL_VOL)
    #                 print(row.WELL)
    #                 print(excelData[row.DATETIME][row.WELL])


def getExcelForDay(date):

    excelData = dict()
    currentDay = fr'\{date.year}\{date.strftime("%m")}-{date.strftime("%b")}\{date.strftime("%d")}'

    for file in os.listdir(path + currentDay):
        if file.endswith(".xlsm") and not file.startswith("~$"):
            print(os.path.join(path + currentDay, file))
            wb = openpyxl.load_workbook(os.path.join(
                path + currentDay, file), data_only=True, read_only=True)
            if "WELL DATA" in wb.sheetnames:
                ws = wb["WELL DATA"]
                for row in ws.iter_rows(min_row=6, min_col=2, values_only=True):
                    if row[0]:
                        key = "BOLT-113-1" if row[0] == "BOLT-113" else row[0]
                        excelData[key.strip()] = {
                            "status": row[2],
                            "hrs_online": row[3],
                            "bopd": row[37],
                            "bsw": row[35],
                            "bfpd": row[34],
                            "rate": (row[3] if row[3] is not None and row[3] != '' else 0) / 24 * row[37],
                            "pIntake": row[26]
                        }
            if "WATER FLOOD WELLS" in wb.sheetnames:
                ws = wb["WATER FLOOD WELLS"]
                for row in ws.iter_rows(min_row=4, min_col=2, values_only=True):
                    if row[0] == "INJECTED WATER SUMMARY" or row[0] == 'WATER FLOOD SOURCE WELLS':
                        break
                    if row[0]:
                        if row[0] in excelData:
                            print(row[0])
                            print("found already")
                            # sys.exit()

                        excelData[row[0].strip()] = {
                            "status": row[2],
                            "hrs_online": row[3],
                            "inj_rate": row[7] if row[7] is not None and row[7] != '' else 0
                        }

    return excelData


main()
